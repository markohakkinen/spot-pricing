from datetime import datetime
from os import PathLike
from typing import AnyStr

from dateutil import relativedelta
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pytz import timezone
from zaptec import UserChargeHistory


class ZaptecInvoice:
    """Create invoice of Zaptec charging based on spot prices"""

    def create(
        self,
        filename: str | PathLike[AnyStr],
        year: int,
        month: int,
        zone: str,
        contract: dict[str, float],
        day_ahead_prices: dict[datetime, float],
        user_charge_histories: dict[str, UserChargeHistory],
    ) -> None:
        (start, end) = self._get_start_end_times(year, month, zone)
        wb = Workbook()
        wb.iso_dates = True
        self._fill_invoising(wb, contract, user_charge_histories)
        self._fill_contract_info(wb, contract)
        self._fill_spot_price(wb, start, end, zone, contract, day_ahead_prices)
        self._fill_charge_histories(wb, start, end, zone, user_charge_histories)
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        wb.save(filename)  # type:ignore

    def _get_start_end_times(
        self, year: int, month: int, zone: str
    ) -> tuple[datetime, datetime]:
        start = datetime(year=year, month=month, day=1)
        end = start + relativedelta.relativedelta(months=1)
        tz = timezone(zone)
        return (tz.localize(start), tz.localize(end))

    def _fill_invoising(
        self,
        wb: Workbook,
        contract: dict[str, float],
        user_charge_histories: dict[str, UserChargeHistory],
    ) -> None:
        invoising = wb.create_sheet("Laskutus")
        invoising.column_dimensions["A"].width = 20
        invoising.column_dimensions["B"].width = 20
        invoising.column_dimensions["C"].width = 16
        invoising.column_dimensions["D"].width = 16
        invoising.column_dimensions["E"].width = 16
        invoising.column_dimensions["F"].width = 20
        invoising.column_dimensions["G"].width = 20
        invoising.column_dimensions["H"].width = 20
        invoising.column_dimensions["I"].width = 20
        invoising.append(
            [
                "Tili",
                "Käyttäjä",
                "Kulutus kWh",
                f"Hinta € ALV {contract['value_added_tax_percentage']}%",
                "Hinta € ALV 0%",
                f"Hinta c/kWh ALV {contract['value_added_tax_percentage']}%",
                "Hinta c/kWh ALV 0%",
                f"Spot hinta c/kWh ALV {contract['value_added_tax_percentage']}%",
                "Spot hinta c/kWh ALV 0%",
            ]
        )
        vat_factor = 1 + contract["value_added_tax_percentage"] / 100
        last_row = 0
        for i, (user, charge_history) in enumerate(user_charge_histories.items()):
            invoising.append(
                [
                    user,
                    charge_history.full_name,
                    f"=SUM('{self._limit_sheet_name(user)}'!B:B)",
                    f"=E{i+2}*{vat_factor}",
                    f"=SUM('{self._limit_sheet_name(user)}'!F:F)",
                    f"=D{i+2}/C{i+2}*100",
                    f"=E{i+2}/C{i+2}*100",
                    f"=F{i+2}-'Sähkösopimus'!B5",
                    f"=G{i+2}-'Sähkösopimus'!C5",
                ]
            )
            last_row = i
        invoising.append(
            [
                "Yhteensä",
                "",
                f"=SUM(C2:C{last_row+2})",
                f"=SUM(D2:D{last_row+2})",
                f"=SUM(E2:E{last_row+2})",
                f"=D{last_row+3}/C{last_row+3}*100",
                f"=E{last_row+3}/C{last_row+3}*100",
                f"=F{last_row+3}-'Sähkösopimus'!B5",
                f"=G{last_row+3}-'Sähkösopimus'!C5",
            ]
        )
        for i in range(2, last_row + 4):
            invoising[f"D{i}"].number_format = "0.00"
            invoising[f"E{i}"].number_format = "0.00"
            invoising[f"F{i}"].number_format = "0.00"
            invoising[f"G{i}"].number_format = "0.00"
            invoising[f"H{i}"].number_format = "0.00"
            invoising[f"I{i}"].number_format = "0.00"

    def _fill_contract_info(
        self,
        wb: Workbook,
        contract: dict[str, float],
    ) -> None:
        vat_factor = 1 + contract["value_added_tax_percentage"] / 100
        contract_info: Worksheet = wb.create_sheet("Sähkösopimus")
        contract_info.column_dimensions["A"].width = 20
        contract_info.column_dimensions["B"].width = 20
        contract_info.column_dimensions["C"].width = 20
        contract_info.append(
            ["", f"c/kWh ALV {contract['value_added_tax_percentage']}%", "c/kWh ALV 0%"]
        )
        contract_info.append(
            [
                "Sähkönsiirto",
                contract["transfer_price_with_vat_c_kwh"],
                f"=B2/{vat_factor}",
            ]
        )
        contract_info.append(
            ["Vero", contract["electricity_tax_with_vat_c_kwh"], f"=B3/{vat_factor}"]
        )
        contract_info.append(
            ["Välityspalkkio", contract["margin_with_vat_c_kwh"], f"=B4/{vat_factor}"]
        )
        contract_info.append(["Kiinteä tuntihinta", "=SUM(B2:B4)", "=SUM(C2:C4)"])

    def _fill_charge_histories(
        self,
        wb: Workbook,
        start: datetime,
        end: datetime,
        zone: str,
        user_charge_histories: dict[str, UserChargeHistory],
    ) -> None:
        for user, charge_history in user_charge_histories.items():
            user_consumption_sheet: Worksheet = wb.create_sheet(
                self._limit_sheet_name(user)
            )
            self._fill_user_consumption_sheet(
                user_consumption_sheet, start, end, zone, charge_history
            )

    def _fill_user_consumption_sheet(
        self,
        user_consumption_sheet: Worksheet,
        start: datetime,
        end: datetime,
        zone: str,
        charge_history: UserChargeHistory,
    ) -> None:
        user_consumption_sheet.column_dimensions["A"].width = 20
        user_consumption_sheet.column_dimensions["B"].width = 20
        user_consumption_sheet.column_dimensions["C"].width = 20
        user_consumption_sheet.column_dimensions["D"].width = 20
        user_consumption_sheet.column_dimensions["E"].width = 20
        user_consumption_sheet.column_dimensions["F"].width = 20
        user_consumption_sheet.append(
            [
                "Aika (UTC)",
                "Kulutus kWh",
                "Spot hinta c/kWh ALV 0%",
                "Spot hinta € ALV 0%",
                "Kiinteä hinta € ALV 0%",
                "Hinta € ALV 0%",
            ]
        )
        sorted_consumption = dict(sorted(charge_history.consumption.items()))
        row = 2
        for key, value in sorted_consumption.items():
            if (key >= start) and (key < end):
                time = key.astimezone(timezone(zone)).replace(tzinfo=None)
                user_consumption_sheet.append(
                    [
                        time,
                        value,
                        f"=VLOOKUP(A{row},'Spot-hinta'!A:C,3,TRUE)",
                        f"=B{row}*C{row}/100",
                        f"='Sähkösopimus'!$C$5*B{row}/100",
                        f"=D{row}+E{row}",
                    ]
                )
                row += 1

    def _fill_spot_price(
        self,
        wb: Workbook,
        start: datetime,
        end: datetime,
        zone: str,
        contract: dict[str, float],
        day_ahead_prices: dict[datetime, float],
    ) -> None:
        vat_factor = 1 + contract["value_added_tax_percentage"] / 100
        spot_price_sheet: Worksheet = wb.create_sheet("Spot-hinta")
        spot_price_sheet.column_dimensions["A"].width = 20
        spot_price_sheet.column_dimensions["B"].width = 16
        spot_price_sheet.column_dimensions["C"].width = 16
        spot_price_sheet.column_dimensions["D"].width = 16
        spot_price_sheet.append(
            [
                f"Aika ({zone})",
                "€/MWh (ALV 0%)",
                "c/kWh (ALV 0%)",
                f"c/kWh (ALV {contract['value_added_tax_percentage']}%)",
            ]
        )
        row = 2
        for key, value in day_ahead_prices.items():
            if (key >= start) and (key < end):
                time = key.astimezone(timezone(zone)).replace(tzinfo=None)
                spot_price_sheet.append(
                    [time, value, f"=B{row}/1000*100", f"=C{row}*{vat_factor}"]
                )
                row += 1

    @staticmethod
    def _limit_sheet_name(name: str) -> str:
        return name[:30]
